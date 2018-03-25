# -*- coding: utf-8 -*-
"""
Created on Tue Dec  5 15:57:31 2017

@author: woon.zhenhao
"""

import os
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd

ordHist='order history'
ordTemplate='sv upload template'
ordUpload='sv upload/SV upload '
invenLogger='inventory updates/inven updates '

def readReport(folderName):
    found = False
    
    for f in os.listdir('./'+folderName):
        if f[:1]!= '~':
            name=f
            found = True
    
    if found==True:
        wbname=folderName+'/'+name
        wb=load_workbook(folderName+'/'+name)
        ws=wb.worksheets[0]
        
        return True, wb, ws, wbname
    
    else:
        return False, None
    
def getOrdList():
    found,wb,ws, wbname = readReport(ordHist)
    
    if found:       
        ls=[]
        
        row=int(ws.max_row)
        if row > 1:
            for r in range(2,row+1):
                ls.append(ws.cell(r,1).internal_value)
        
    return found,ls

#ls=getOrdList()

def writeToMaster(ls):
    found, wb, ws, wbname = readReport(ordHist)
    
    if found:
        row=int(ws.max_row)+1
        
        for i in ls.index:
            r=ls.loc[i]
            ws.cell(row=row,column=1).value = r[0]
            ws.cell(row=row,column=2).value = r[1]
            ws.cell(row=row,column=3).value = r[2]
            ws.cell(row=row,column=4).value = r[3]
            row+=1        
    wb.save(wbname)
    
def writeToTemplate(df, account):
    found, wb, ws, wbname = readReport(ordTemplate)
    
    header = next(ws.values)[0:]
    
    df.columns = header
        
    name = ordUpload+account+' '+datetime.now().strftime("%d-%m-%y_%H-%M-%S")+'.csv'
    df.to_csv(name, header=True, index=False)
        